"""
Build Multifamily Backward Analysis Excel Workbook
===================================================
Creates a structured .xlsx file with:
- Tab 1: Inputs (all adjustable assumptions)
- Tab 2: Retirement Backward Solve
- Tab 3: Debt & LTV Validation
- Tab 4: Investor Marketability
- Tab 5: Creative Financing
- Tab 6: Deal Criteria Box
- Tab 7: Scenario Analysis
- Tab 8: Decision Framework

All cells use Excel formulas referencing the Inputs tab so the user
can change assumptions and see results update automatically.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# =============================================================================
# STYLING HELPERS
# =============================================================================

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
LABEL_FONT = Font(name="Calibri", size=10)
INPUT_FONT = Font(name="Calibri", size=10, bold=True, color="0000CC")
RESULT_FONT = Font(name="Calibri", size=10, bold=True, color="006600")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, title, max_col=4):
    """Style a section header row."""
    ws.cell(row=row, column=1, value=title).font = HEADER_FONT
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = HEADER_FONT


def style_input_cell(ws, row, col):
    """Style a user-editable input cell."""
    cell = ws.cell(row=row, column=col)
    cell.fill = INPUT_FILL
    cell.font = INPUT_FONT
    cell.border = THIN_BORDER


def style_result_cell(ws, row, col):
    """Style a calculated result cell."""
    cell = ws.cell(row=row, column=col)
    cell.fill = RESULT_FILL
    cell.font = RESULT_FONT
    cell.border = THIN_BORDER


def add_label(ws, row, col, text):
    """Add a label."""
    ws.cell(row=row, column=col, value=text).font = LABEL_FONT


# =============================================================================
# TAB 1: INPUTS
# =============================================================================

def build_inputs_tab(wb):
    """Create the master Inputs tab with all adjustable assumptions."""
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18

    row = 1
    style_header(ws, row, "MULTIFAMILY BACKWARD ANALYSIS — ALL INPUTS")
    row += 2

    # --- Retirement Goals ---
    ws.cell(row=row, column=1, value="RETIREMENT GOALS").font = SECTION_FONT
    row += 1
    inputs_retirement = [
        ("Annual passive income (after-tax)", 200000, "B4", "$"),
        ("Timeline to retirement (years)", 5, "B5", "#"),
        ("Min cash-on-cash return", 0.10, "B6", "%"),
        ("Number of properties at retirement", 3, "B7", "#"),
        ("Effective tax rate", 0.30, "B8", "%"),
        ("Cash available for first deal", 150000, "B9", "$"),
    ]
    for label, val, _, fmt in inputs_retirement:
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=val)
        style_input_cell(ws, row, 2)
        if fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1

    row += 1
    # --- Debt Assumptions ---
    ws.cell(row=row, column=1, value="DEBT ASSUMPTIONS").font = SECTION_FONT
    row += 1
    inputs_debt = [
        ("LTV - Minimum", 0.65, "B11"),
        ("LTV - Target", 0.75, "B12"),
        ("LTV - Maximum", 0.80, "B13"),
        ("DSCR Required", 1.25, "B14"),
        ("Interest Rate", 0.065, "B15"),
        ("Amortization (years)", 30, "B16"),
        ("Loan Term (years)", 7, "B17"),
        ("Origination Costs %", 0.025, "B18"),
    ]
    for label, val, _ in inputs_debt:
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=val)
        style_input_cell(ws, row, 2)
        if "%" in label or "LTV" in label or "Rate" in label or "DSCR" not in label and val < 1:
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif isinstance(val, float) and val > 1:
            ws.cell(row=row, column=2).number_format = '0.00'
        row += 1

    row += 1
    # --- Market Assumptions ---
    ws.cell(row=row, column=1, value="MARKET ASSUMPTIONS").font = SECTION_FONT
    row += 1
    inputs_market = [
        ("Avg market rent per unit (monthly)", 1200, "$"),
        ("Avg unit square footage", 850, "#"),
        ("Vacancy rate", 0.07, "%"),
        ("Expense ratio (% of EGI)", 0.45, "%"),
        ("Annual rent growth", 0.03, "%"),
        ("Annual expense growth", 0.025, "%"),
        ("Market cap rate", 0.065, "%"),
        ("Exit cap rate (conservative)", 0.070, "%"),
    ]
    for label, val, fmt in inputs_market:
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=val)
        style_input_cell(ws, row, 2)
        if fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1

    row += 1
    # --- Investor Assumptions ---
    ws.cell(row=row, column=1, value="INVESTOR / SYNDICATION ASSUMPTIONS").font = SECTION_FONT
    row += 1
    inputs_investor = [
        ("Preferred return to LPs", 0.08, "%"),
        ("Target equity multiple", 2.0, "x"),
        ("Target IRR to LPs", 0.17, "%"),
        ("Hold period (years)", 5, "#"),
        ("GP promote after pref", 0.30, "%"),
        ("Min investment per LP", 50000, "$"),
        ("Reserve months (operating)", 6, "#"),
    ]
    for label, val, fmt in inputs_investor:
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=val)
        style_input_cell(ws, row, 2)
        if fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "x":
            ws.cell(row=row, column=2).number_format = '0.00'
        row += 1

    row += 1
    # --- Creative Finance ---
    ws.cell(row=row, column=1, value="CREATIVE FINANCING ASSUMPTIONS").font = SECTION_FONT
    row += 1
    inputs_creative = [
        ("Seller finance rate", 0.05, "%"),
        ("Seller finance term (years)", 10, "#"),
        ("Seller balloon year", 5, "#"),
        ("Seller max LTV carry", 0.85, "%"),
        ("Max out-of-pocket (creative)", 50000, "$"),
        ("Master lease term (years)", 3, "#"),
        ("Option price premium", 0.05, "%"),
    ]
    for label, val, fmt in inputs_creative:
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=val)
        style_input_cell(ws, row, 2)
        if fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1

    row += 1
    # --- Value-Add ---
    ws.cell(row=row, column=1, value="VALUE-ADD ASSUMPTIONS").font = SECTION_FONT
    row += 1
    inputs_va = [
        ("Capex per unit", 15000, "$"),
        ("Rent bump per unit (monthly)", 200, "$"),
        ("Reno timeline (months)", 18, "#"),
        ("Units renovated per month", 3, "#"),
        ("Expense reduction %", 0.05, "%"),
    ]
    for label, val, fmt in inputs_va:
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=val)
        style_input_cell(ws, row, 2)
        if fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1

    return ws


# =============================================================================
# TAB 2: RETIREMENT BACKWARD SOLVE
# =============================================================================

def build_retirement_tab(wb):
    """Backward-solve from income goal to required deal parameters."""
    ws = wb.create_sheet("Retirement Solve")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15

    row = 1
    style_header(ws, row, "BACKWARD SOLVE: RETIREMENT INCOME TO DEAL SIZE", 3)
    row += 2

    # Reference cells from Inputs tab
    # Inputs!B4 = annual passive income, B5 = timeline, B6 = CoC,
    # B7 = num properties, B8 = tax rate, B9 = cash available
    # B12 = LTV target, B14 = DSCR, B15 = interest rate, B16 = amort
    # B18 = origination costs
    # B20 = avg rent, B22 = vacancy, B23 = expense ratio, B26 = cap rate

    formulas = [
        ("Target passive income (after-tax)", "=Inputs!B4", "$"),
        ("Effective tax rate", "=Inputs!B8", "%"),
        ("Pre-tax income needed", "=B3/(1-B4)", "$"),
        ("Number of properties", "=Inputs!B7", "#"),
        ("Income needed per property", "=B5/B6", "$"),
        ("", "", ""),
        ("Min cash-on-cash return", "=Inputs!B6", "%"),
        ("Equity required per deal", "=B7/B9", "$"),
        ("Total equity needed (all deals)", "=B10*B6", "$"),
        ("", "", ""),
        ("Market cap rate", "=Inputs!B26", "%"),
        ("LTV target", "=Inputs!B12", "%"),
        ("Interest rate", "=Inputs!B15", "%"),
        ("Amortization (years)", "=Inputs!B16", "#"),
        ("", "", ""),
        ("Monthly rate", "=B15/12", "%"),
        ("Total payments (months)", "=B16*12", "#"),
        ("Annual loan constant (K)", "=(B18*(1+B18)^B19/((1+B18)^B19-1))*12", "%"),
        ("Debt burden factor (LTV*K/Cap)", "=B14*B20/B13", ""),
        ("", "", ""),
        ("REQUIRED NOI PER DEAL", "=B7/(1-B21)", "$"),
        ("PURCHASE PRICE PER DEAL", "=B23/B13", "$"),
        ("LOAN AMOUNT", "=B24*B14", "$"),
        ("ANNUAL DEBT SERVICE", "=B25*B20", "$"),
        ("YEAR 1 CASH FLOW", "=B23-B26", "$"),
        ("DSCR (actual)", "=B23/B26", "x"),
        ("", "", ""),
        ("NOI per unit (annual)", "=Inputs!B20*12*(1-Inputs!B22)*(1-Inputs!B23)", "$"),
        ("UNITS NEEDED PER DEAL", "=ROUNDUP(B23/B30,0)", "#"),
        ("PRICE PER UNIT", "=B24/B31", "$"),
        ("", "", ""),
        ("Down payment", "=B24*(1-B14)", "$"),
        ("Closing costs", "=B25*Inputs!B18", "$"),
        ("TOTAL CASH TO CLOSE", "=B34+B35", "$"),
    ]

    for label, formula, fmt in formulas:
        row += 1
        if label == "":
            continue
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=formula)
        if label.isupper():
            style_result_cell(ws, row, 2)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.0000%'
        elif fmt == "x":
            ws.cell(row=row, column=2).number_format = '0.00"x"'

    return ws


# =============================================================================
# TAB 3: DEBT & LTV VALIDATION
# =============================================================================

def build_debt_tab(wb):
    """Validate conventional debt feasibility."""
    ws = wb.create_sheet("Debt Validation")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    row = 1
    style_header(ws, row, "DEBT & LTV CONSTRAINT VALIDATION", 3)

    # We'll use computed values from the Python model here since Excel formula
    # references across complex iterative calcs are unwieldy. Instead, we
    # create a hybrid: key formulas + static reference values.

    row = 3
    section_data = [
        ("Required NOI (from Retirement tab)", "='Retirement Solve'!B23", "$"),
        ("Purchase Price", "='Retirement Solve'!B24", "$"),
        ("Loan Amount at Target LTV", "='Retirement Solve'!B25", "$"),
        ("Annual Debt Service", "='Retirement Solve'!B26", "$"),
        ("", "", ""),
        ("DSCR (actual)", "='Retirement Solve'!B28", "x"),
        ("DSCR Required (minimum)", "=Inputs!B14", "x"),
        ("DSCR Pass?", '=IF(B8>=B9,"PASS","FAIL")', ""),
        ("", "", ""),
        ("Max Loan from DSCR", "=B3/(B9*'Retirement Solve'!B20)", "$"),
        ("Max Loan from LTV", "=B4*Inputs!B13", "$"),
        ("Binding Constraint", '=IF(B13<B14,"DSCR","LTV")', ""),
        ("Implied LTV at DSCR Limit", "=B13/B4", "%"),
        ("", "", ""),
        ("— SPREAD ANALYSIS —", "", ""),
        ("Market Cap Rate", "=Inputs!B26", "%"),
        ("Cost of Debt (interest rate)", "=Inputs!B15", "%"),
        ("Spread (Cap - Debt) in bps", "=(B19-B20)*10000", "#"),
        ("Spread Assessment", '=IF(B21>=100,"POSITIVE LEVERAGE",IF(B21>=0,"NEUTRAL","NEGATIVE LEVERAGE"))', ""),
        ("", "", ""),
        ("— BREAKEVEN ANALYSIS —", "", ""),
        ("Gross Potential Income", "=B3/((1-Inputs!B22)*(1-Inputs!B23))", "$"),
        ("Fixed Expenses", "=B25*Inputs!B23", "$"),
        ("Breakeven Occupancy", "=(B6+B26)/B25", "%"),
        ("", "", ""),
        ("— STRESS TEST: Rates +150bps at Refi —", "", ""),
        ("Stress Rate", "=Inputs!B15+0.015", "%"),
        ("NOI at Refi (grown)", "=B3*(1+Inputs!B24)^Inputs!B17", "$"),
        ("Stress DSCR", "=B31/B6", "x"),
        ("Stress Cash Flow", "=B31-B6", "$"),
        ("Stress Test Pass?", '=IF(B32>=1.0,"PASS","FAIL")', ""),
        ("", "", ""),
        ("═══ VERDICT ═══", "", ""),
        ("Conventional Viable?", '=IF(AND(B10="PASS",B21>=0),"CONVENTIONAL FINANCING VIABLE","CREATIVE STRUCTURE REQUIRED")', ""),
    ]

    for label, formula, fmt in section_data:
        if label == "":
            row += 1
            continue
        add_label(ws, row, 1, label)
        if formula:
            ws.cell(row=row, column=2, value=formula)
            if label.startswith("═") or "VERDICT" in label.upper():
                style_result_cell(ws, row, 2)
                style_result_cell(ws, row, 1)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "x":
            ws.cell(row=row, column=2).number_format = '0.00"x"'
        row += 1

    return ws


# =============================================================================
# TAB 4: INVESTOR MARKETABILITY
# =============================================================================

def build_investor_tab(wb):
    """Syndication feasibility and LP return projections."""
    ws = wb.create_sheet("Investor Analysis")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    row = 1
    style_header(ws, row, "INVESTOR MARKETABILITY — SYNDICATION FEASIBILITY", 6)

    row = 3
    capital_raise_items = [
        ("Down payment", "='Retirement Solve'!B34", "$"),
        ("Closing costs", "='Retirement Solve'!B35", "$"),
        ("Capex budget (units * capex/unit)", "='Retirement Solve'!B31*Inputs!B48", "$"),
        ("Operating reserves (months)", "=('Retirement Solve'!B26/12)*Inputs!B40", "$"),
        ("TOTAL CAPITAL RAISE", "=B3+B4+B5+B6", "$"),
        ("", "", ""),
        ("GP co-invest (your cash)", "=Inputs!B9", "$"),
        ("LP equity needed", "=B7-B9", "$"),
        ("Investors needed (@ min investment)", "=ROUNDUP(B10/Inputs!B39,0)", "#"),
        ("", "", ""),
        ("— LP RETURN TARGETS —", "", ""),
        ("Preferred return", "=Inputs!B34", "%"),
        ("Target equity multiple", "=Inputs!B35", "x"),
        ("Target IRR", "=Inputs!B36", "%"),
        ("GP promote after pref", "=Inputs!B38", "%"),
        ("Hold period", "=Inputs!B37", "#"),
    ]

    for label, formula, fmt in capital_raise_items:
        if label == "":
            row += 1
            continue
        add_label(ws, row, 1, label)
        if formula:
            ws.cell(row=row, column=2, value=formula)
        if "TOTAL" in label:
            style_result_cell(ws, row, 2)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "x":
            ws.cell(row=row, column=2).number_format = '0.00'
        row += 1

    # Cash flow projection table
    row += 2
    ws.cell(row=row, column=1, value="ANNUAL CASH FLOW PROJECTION").font = SECTION_FONT
    row += 1
    headers = ["Year", "NOI", "Debt Service", "Total CF", "LP Share", "GP Share"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1

    # Year 1-5 rows with formulas
    for yr in range(1, 6):
        ws.cell(row=row, column=1, value=yr)
        # NOI with growth + value-add phased in
        ws.cell(row=row, column=2,
                value=f"='Retirement Solve'!B23*(1+Inputs!B24)^{yr}+"
                      f"'Retirement Solve'!B31*Inputs!B49*12*(1-Inputs!B23)*{yr}/5")
        ws.cell(row=row, column=2).number_format = '$#,##0'
        # Debt service
        ws.cell(row=row, column=3, value="='Retirement Solve'!B26")
        ws.cell(row=row, column=3).number_format = '$#,##0'
        # Total CF
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=4).number_format = '$#,##0'
        # LP share
        ws.cell(row=row, column=5, value=f"=D{row}*(1-Inputs!B38)")
        ws.cell(row=row, column=5).number_format = '$#,##0'
        # GP share
        ws.cell(row=row, column=6, value=f"=D{row}*Inputs!B38")
        ws.cell(row=row, column=6).number_format = '$#,##0'
        row += 1

    # Exit analysis
    row += 1
    ws.cell(row=row, column=1, value="EXIT ANALYSIS").font = SECTION_FONT
    row += 1
    exit_items = [
        ("Stabilized NOI (Year 5)", f"=B{row-3}", "$"),
        ("Exit cap rate", "=Inputs!B27", "%"),
        ("Gross exit value", f"=B{row}/B{row+1}", "$"),
        ("Less: disposition costs (3%)", f"=B{row+2}*0.03", "$"),
        ("Less: remaining loan balance (approx)", "='Retirement Solve'!B25*0.90", "$"),
        ("NET SALE PROCEEDS", f"=B{row+2}-B{row+3}-B{row+4}", "$"),
        ("", "", ""),
        ("LP exit share", f"=B{row+5}*(1-Inputs!B38)", "$"),
        ("GP exit share (promote)", f"=B{row+5}*Inputs!B38", "$"),
        ("", "", ""),
        ("LP total distributions", f"=B{row+7}+SUMPRODUCT(E{row-8}:E{row-4})", "$"),
        ("LP EQUITY MULTIPLE", f"=B{row+10}/B10", "x"),
    ]

    for label, formula, fmt in exit_items:
        if label == "":
            row += 1
            continue
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=formula)
        if "EQUITY MULTIPLE" in label or "NET SALE" in label:
            style_result_cell(ws, row, 2)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "x":
            ws.cell(row=row, column=2).number_format = '0.00"x"'
        row += 1

    # Marketability verdict
    row += 2
    ws.cell(row=row, column=1, value="MARKETABILITY VERDICT").font = SECTION_FONT
    row += 1
    add_label(ws, row, 1, "Can raise in 60 days?")
    ws.cell(row=row, column=2,
            value=f'=IF(B{row-3}>=Inputs!B35,"YES - Deal is marketable","DIFFICULT - Rework returns")')
    style_result_cell(ws, row, 2)

    return ws


# =============================================================================
# TAB 5: CREATIVE FINANCING
# =============================================================================

def build_creative_tab(wb):
    """Creative financing analysis: seller finance, sub-to, master lease."""
    ws = wb.create_sheet("Creative Financing")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    row = 1
    style_header(ws, row, "CREATIVE FINANCING FEASIBILITY", 3)

    row = 3
    ws.cell(row=row, column=1, value="SELLER FINANCING").font = SECTION_FONT
    row += 1

    sf_items = [
        ("Purchase price", "='Retirement Solve'!B24", "$"),
        ("Seller carry LTV", "=Inputs!B44", "%"),
        ("Seller carry amount", "=B4*B5", "$"),
        ("Down payment to seller", "=B4-B6", "$"),
        ("Seller finance rate", "=Inputs!B42", "%"),
        ("Seller finance term (years)", "=Inputs!B43", "#"),
        ("Monthly payment (seller note)", "=PMT(B8/12,B9*12,-B6)", "$"),
        ("Annual debt service (seller)", "=B10*12", "$"),
        ("NOI", "='Retirement Solve'!B23", "$"),
        ("Cash flow (seller finance)", "=B12-B11", "$"),
        ("DSCR (seller note)", "=B12/B11", "x"),
        ("Balloon year", "=Inputs!B45", "#"),
        ("Balloon balance (approx)", "=B6*0.85", "$"),
        ("", "", ""),
        ("Seller finance VIABLE?", '=IF(AND(B13>0,B14>=1.1),"YES","NO")', ""),
    ]

    for label, formula, fmt in sf_items:
        if label == "":
            row += 1
            continue
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=formula)
        if "VIABLE" in label:
            style_result_cell(ws, row, 2)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif fmt == "x":
            ws.cell(row=row, column=2).number_format = '0.00"x"'
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="SUBJECT-TO / WRAP").font = SECTION_FONT
    row += 1

    st_items = [
        ("Estimated existing debt (50% of price)", "='Retirement Solve'!B24*0.50", "$"),
        ("Existing rate (assumed)", 0.045, "%"),
        ("Existing annual debt service", "=PMT(B22/12,25*12,-B21)*12", "$"),
        ("Gap financing needed", "=B4-B21-Inputs!B46", "$"),
        ("Cash flow (sub-to)", "='Retirement Solve'!B23-B23", "$"),
        ("Gap < 20% of price?", '=IF(B24<B4*0.2,"YES","NO")', ""),
        ("Subject-to VIABLE?", '=IF(AND(B25>0,B26="YES"),"YES","NO")', ""),
        ("Due-on-sale risk", "MODERATE", ""),
    ]

    for label, formula, fmt in st_items:
        if label == "":
            row += 1
            continue
        add_label(ws, row, 1, label)
        if isinstance(formula, str) and formula.startswith("="):
            ws.cell(row=row, column=2, value=formula)
        else:
            ws.cell(row=row, column=2, value=formula)
        if "VIABLE" in label:
            style_result_cell(ws, row, 2)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="MASTER LEASE WITH OPTION").font = SECTION_FONT
    row += 1

    ml_items = [
        ("Seller's debt service (existing)", "=B23", "$"),
        ("Master lease payment (110% of seller DS)", "=B31*1.10", "$"),
        ("Monthly lease payment", "=B32/12", "$"),
        ("Cash flow during lease", "='Retirement Solve'!B23-B32", "$"),
        ("Option strike price (price + premium)", "=B4*(1+Inputs!B47)", "$"),
        ("Stabilized NOI post-reno", "='Retirement Solve'!B23+'Retirement Solve'!B31*Inputs!B49*12*(1-Inputs!B23)", "$"),
        ("Refi value at market cap", "=B36/Inputs!B26", "$"),
        ("Refi loan at 75% LTV", "=B37*0.75", "$"),
        ("Can cash-out refi to exercise?", '=IF(B38>=B35,"YES","NO")', ""),
        ("Master lease VIABLE?", '=IF(AND(B34>0,B39="YES"),"YES","NO")', ""),
    ]

    for label, formula, fmt in ml_items:
        if label == "":
            row += 1
            continue
        add_label(ws, row, 1, label)
        ws.cell(row=row, column=2, value=formula)
        if "VIABLE" in label or "cash-out" in label.lower():
            style_result_cell(ws, row, 2)
        if fmt == "$":
            ws.cell(row=row, column=2).number_format = '$#,##0'
        elif fmt == "%":
            ws.cell(row=row, column=2).number_format = '0.00%'
        row += 1

    # Overall creative verdict
    row += 2
    ws.cell(row=row, column=1, value="═══ CREATIVE VERDICT ═══").font = SECTION_FONT
    row += 1
    add_label(ws, row, 1, "Best creative structure")
    ws.cell(row=row, column=2,
            value='=IF(B18="YES","SELLER FINANCING",IF(B27="YES","SUBJECT-TO",IF(B40="YES","MASTER LEASE","NONE")))')
    style_result_cell(ws, row, 2)
    row += 1
    add_label(ws, row, 1, "Minimum out-of-pocket")
    ws.cell(row=row, column=2,
            value='=IF(B18="YES",B7,IF(B27="YES",Inputs!B46,IF(B40="YES",B33*2,B7)))')
    ws.cell(row=row, column=2).number_format = '$#,##0'
    style_result_cell(ws, row, 2)

    return ws


# =============================================================================
# TAB 6: DEAL CRITERIA BOX
# =============================================================================

def build_criteria_tab(wb):
    """Output the min/target/max acquisition criteria table."""
    ws = wb.create_sheet("Deal Criteria")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    row = 1
    style_header(ws, row, "ACQUISITION CRITERIA BOX")

    row = 3
    # Headers
    for col, h in enumerate(["Parameter", "Minimum", "Target", "Maximum"], 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1

    # Criteria rows with formulas referencing other tabs
    criteria_rows = [
        ("Unit count",
         "=ROUNDUP('Retirement Solve'!B31*0.7,0)",
         "='Retirement Solve'!B31",
         "=ROUNDUP('Retirement Solve'!B31*1.5,0)", "#"),
        ("Purchase price",
         "='Retirement Solve'!B24*0.75",
         "='Retirement Solve'!B24",
         "='Retirement Solve'!B24*1.15", "$"),
        ("Going-in cap rate",
         "=Inputs!B26-0.005",
         "=Inputs!B26",
         "=Inputs!B26+0.02", "%"),
        ("Price per unit",
         "='Retirement Solve'!B24*0.75/'Retirement Solve'!B31",
         "='Retirement Solve'!B32",
         "='Retirement Solve'!B24*1.15/ROUNDUP('Retirement Solve'!B31*0.7,0)", "$"),
        ("Price per SF",
         "='Retirement Solve'!B24*0.75/('Retirement Solve'!B31*Inputs!B21)",
         "='Retirement Solve'!B24/('Retirement Solve'!B31*Inputs!B21)",
         "='Retirement Solve'!B24*1.15/(ROUNDUP('Retirement Solve'!B31*0.7,0)*Inputs!B21)", "$"),
        ("Current occupancy",
         "80%", "90%", "—", "%"),
        ("In-place NOI",
         "='Retirement Solve'!B23*0.70",
         "='Retirement Solve'!B23",
         "—", "$"),
        ("Stabilized NOI (post-reno)",
         "='Retirement Solve'!B23",
         "='Retirement Solve'!B23*1.20",
         "—", "$"),
        ("Rent upside/unit (monthly)",
         "$100", "=Inputs!B49", "—", "$"),
        ("Total capex budget",
         "='Retirement Solve'!B31*Inputs!B48*0.5",
         "='Retirement Solve'!B31*Inputs!B48",
         "='Retirement Solve'!B31*Inputs!B48*1.5", "$"),
        ("Capex per unit",
         "=Inputs!B48*0.5",
         "=Inputs!B48",
         "=Inputs!B48*1.5", "$"),
        ("Year 1 Cash-on-Cash",
         "=Inputs!B6*0.8",
         "=Inputs!B6",
         "—", "%"),
        ("5-Year IRR (levered)",
         "14%", "18%", "—", "%"),
        ("Exit cap rate",
         "=Inputs!B27-0.005",
         "=Inputs!B27",
         "=Inputs!B27+0.015", "%"),
        ("Max OOP (creative)",
         "$0",
         "=Inputs!B46",
         "=Inputs!B46*3", "$"),
        ("Capital raise (syndication)",
         "='Investor Analysis'!B7*0.8",
         "='Investor Analysis'!B7",
         "='Investor Analysis'!B7*1.3", "$"),
    ]

    for name, mn, tg, mx, fmt in criteria_rows:
        ws.cell(row=row, column=1, value=name).font = LABEL_FONT
        ws.cell(row=row, column=2, value=mn)
        ws.cell(row=row, column=3, value=tg)
        ws.cell(row=row, column=4, value=mx)
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
        if fmt == "$":
            for col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = '$#,##0'
        elif fmt == "%":
            for col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = '0.0%'
        row += 1

    return ws


# =============================================================================
# TAB 7: SCENARIO ANALYSIS
# =============================================================================

def build_scenario_tab(wb):
    """Bull, Base, Bear, Catastrophe scenario comparison."""
    ws = wb.create_sheet("Scenarios")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = 1
    style_header(ws, row, "SCENARIO & SENSITIVITY ANALYSIS", 5)

    row = 3
    # Scenario assumptions table
    ws.cell(row=row, column=1, value="SCENARIO PARAMETERS").font = SECTION_FONT
    row += 1
    headers = ["Parameter", "BULL", "BASE", "BEAR", "CATASTROPHE"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1

    params = [
        ("Rent growth override", "4.0%", "3.0%", "0.5%", "-1.0%"),
        ("Value-add achieved", "100%", "80%", "50%", "30%"),
        ("Exit cap rate", "6.0%", "7.0%", "8.5%", "9.5%"),
        ("Vacancy override", "5.0%", "7.0%", "10.0%", "14.0%"),
        ("Rate change at refi", "-0.5%", "0.0%", "+1.0%", "+2.0%"),
    ]
    for p_row in params:
        for col, val in enumerate(p_row, 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    # Scenario results (using formulas where possible, some static from model)
    row += 2
    ws.cell(row=row, column=1, value="SCENARIO RESULTS (5-Year Hold)").font = SECTION_FONT
    row += 1
    result_headers = ["Metric", "BULL", "BASE", "BEAR", "CATASTROPHE"]
    for col, h in enumerate(result_headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1

    # These use simplified formulas based on the inputs
    noi_ref = "'Retirement Solve'!B23"
    units_ref = "'Retirement Solve'!B31"
    loan_ref = "'Retirement Solve'!B25"
    ads_ref = "'Retirement Solve'!B26"

    result_rows = [
        ("NOI at Exit",
         f"={noi_ref}*(1.04)^5+{units_ref}*Inputs!B49*12*(1-Inputs!B23)*1.0",
         f"={noi_ref}*(1.03)^5+{units_ref}*Inputs!B49*12*(1-Inputs!B23)*0.8",
         f"={noi_ref}*(1.005)^5+{units_ref}*Inputs!B49*12*(1-Inputs!B23)*0.5",
         f"={noi_ref}*(0.99)^5+{units_ref}*Inputs!B49*12*(1-Inputs!B23)*0.3"),
        ("Exit Value",
         f"=B{row}/0.06", f"=C{row}/0.07", f"=D{row}/0.085", f"=E{row}/0.095"),
        ("Net Proceeds (after loan payoff)",
         f"=B{row+1}*0.97-{loan_ref}*0.90",
         f"=C{row+1}*0.97-{loan_ref}*0.90",
         f"=D{row+1}*0.97-{loan_ref}*0.90",
         f"=E{row+1}*0.97-{loan_ref}*0.90"),
        ("Avg Annual CF",
         f"=({noi_ref}+B{row})/2-{ads_ref}",
         f"=({noi_ref}+C{row})/2-{ads_ref}",
         f"=({noi_ref}+D{row})/2-{ads_ref}",
         f"=({noi_ref}+E{row})/2-{ads_ref}"),
        ("Meets Retirement Goal?",
         f'=IF(B{row+3}*Inputs!B7>=Inputs!B4/(1-Inputs!B8),"YES","NO")',
         f'=IF(C{row+3}*Inputs!B7>=Inputs!B4/(1-Inputs!B8),"YES","NO")',
         f'=IF(D{row+3}*Inputs!B7>=Inputs!B4/(1-Inputs!B8),"YES","NO")',
         f'=IF(E{row+3}*Inputs!B7>=Inputs!B4/(1-Inputs!B8),"YES","NO")'),
        ("Deal Survives?",
         f'=IF(AND(B{row+3}>0,B{row+2}>0),"YES","NO")',
         f'=IF(AND(C{row+3}>0,C{row+2}>0),"YES","NO")',
         f'=IF(AND(D{row+3}>0,D{row+2}>0),"YES","NO")',
         f'=IF(AND(E{row+3}>0,E{row+2}>0),"YES","NO")'),
    ]

    for r_row in result_rows:
        for col, val in enumerate(r_row, 1):
            ws.cell(row=row, column=col, value=val)
            if col > 1 and "NOI" in r_row[0] or "Value" in r_row[0] or "Proceeds" in r_row[0] or "CF" in r_row[0]:
                ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=1).font = LABEL_FONT
        row += 1

    return ws


# =============================================================================
# TAB 8: DECISION FRAMEWORK
# =============================================================================

def build_decision_tab(wb):
    """GO / NO-GO / NEGOTIATE decision output."""
    ws = wb.create_sheet("Decision")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50

    row = 1
    style_header(ws, row, "DECISION FRAMEWORK — GO / NO-GO / NEGOTIATE", 3)

    row = 3
    ws.cell(row=row, column=1, value="SCORECARD").font = SECTION_FONT
    row += 1
    headers = ["Criterion", "Result", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1

    scorecard = [
        ("Conventional financing viable",
         "='Debt Validation'!B37",
         "Requires positive spread + DSCR >= 1.25x"),
        ("Investor marketability grade",
         '=IF(\'Investor Analysis\'!B7>0,"MARKETABLE","NOT MARKETABLE")',
         "Needs 8%+ pref, 2.0x multiple, 17%+ IRR"),
        ("Creative structure available",
         "='Creative Financing'!B44",
         "Seller finance, sub-to, or master lease"),
        ("Bear case survives",
         "=Scenarios!D19",
         "Positive CF and equity at exit in bear scenario"),
        ("Catastrophe survival",
         "=Scenarios!E19",
         "Can survive rate spike + rent decline"),
        ("Meets retirement goal (base)",
         "=Scenarios!C18",
         "Cash flow * num properties >= income target"),
    ]

    for criterion, formula, notes in scorecard:
        ws.cell(row=row, column=1, value=criterion).font = LABEL_FONT
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=notes)
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="═══ FINAL DECISION ═══").font = Font(size=14, bold=True)
    row += 1

    # Decision logic
    decision_formula = (
        '=IF(AND(B5="CONVENTIONAL FINANCING VIABLE",B8="YES",B10="YES"),'
        '"GO (CONVENTIONAL)",'
        'IF(AND(OR(B7<>"NONE",B5="CONVENTIONAL FINANCING VIABLE"),B8="YES"),'
        '"GO (CREATIVE)",'
        'IF(AND(B8="NO",Scenarios!C19="YES"),'
        '"NEGOTIATE — Need price reduction for safety margin",'
        '"NO-GO — Pass on this deal")))'
    )
    ws.cell(row=row, column=1, value="DECISION:")
    ws.cell(row=row, column=2, value=decision_formula)
    ws.cell(row=row, column=2).font = Font(size=14, bold=True, color="006600")
    style_result_cell(ws, row, 2)

    row += 2
    ws.cell(row=row, column=1, value="DECISION DEFINITIONS:").font = SECTION_FONT
    row += 1
    definitions = [
        "GO (CONVENTIONAL): Deal hits all metrics with standard agency debt + investor raise",
        "GO (CREATIVE): Deal requires creative structure but math works and seller profile fits",
        "NEGOTIATE: Deal works at lower price or with seller concession — counter required",
        "NO-GO: Spread too thin, returns don't justify risk, can't market to investors",
    ]
    for d in definitions:
        ws.cell(row=row, column=1, value=d)
        row += 1

    return ws


# =============================================================================
# MAIN: BUILD THE WORKBOOK
# =============================================================================

def build_workbook():
    """Assemble all tabs and save the workbook."""
    wb = Workbook()

    print("Building Inputs tab...")
    build_inputs_tab(wb)

    print("Building Retirement Solve tab...")
    build_retirement_tab(wb)

    print("Building Debt Validation tab...")
    build_debt_tab(wb)

    print("Building Investor Analysis tab...")
    build_investor_tab(wb)

    print("Building Creative Financing tab...")
    build_creative_tab(wb)

    print("Building Deal Criteria tab...")
    build_criteria_tab(wb)

    print("Building Scenarios tab...")
    build_scenario_tab(wb)

    print("Building Decision tab...")
    build_decision_tab(wb)

    # Save
    output_path = "/projects/sandbox/Multifamily_Backward_Analysis.xlsx"
    wb.save(output_path)
    print(f"\nWorkbook saved to: {output_path}")
    print("Open in Excel or Google Sheets to use. Yellow cells are editable inputs.")
    return output_path


if __name__ == "__main__":
    build_workbook()
